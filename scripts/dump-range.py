import openpyxl
from pathlib import Path

ws = openpyxl.load_workbook(
    Path(__file__).parent.parent / "family-data-drive.xlsx", data_only=True
)["All Families"]

empty_run = 0
for r in range(218, 285):
    name = ws.cell(r, 2).value
    if not name:
        empty_run += 1
        print(f"  [{r}] (empty, run={empty_run})")
        continue
    gap = empty_run
    empty_run = 0
    coords = ws.cell(r, 4).value
    addr = (str(ws.cell(r, 5).value or ""))[:50]
    print(f"{r:4d} gap={gap} | {str(name)[:45]:45} | c={bool(coords)} | {addr}")
