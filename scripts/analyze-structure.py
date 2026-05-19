import openpyxl
from pathlib import Path

XLSX = Path(__file__).parent.parent / "family-data-drive.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["All Families"]


def parse_coords(value):
    if not value:
        return None
    text = str(value).strip().strip("'")
    if "," not in text:
        return None
    a, b = text.split(",", 1)
    try:
        return float(a.strip()), float(b.strip())
    except ValueError:
        return None


blocks = []
empty_run = 0
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name:
        empty_run += 1
        continue
    blocks.append(
        {
            "row": r,
            "gap": empty_run,
            "name": str(name).strip(),
            "coords": parse_coords(ws.cell(r, 4).value),
            "addr": str(ws.cell(r, 5).value or "").strip(),
        }
    )
    empty_run = 0

print(f"Total people rows: {len(blocks)}")
print("\n--- Rows with gap >= 1 (family boundaries) ---")
for b in blocks:
    if b["gap"] >= 1:
        print(
            f"row {b['row']:4d} gap={b['gap']} "
            f"coords={bool(b['coords'])} addr={bool(b['addr'])} | {b['name'][:50]}"
        )
