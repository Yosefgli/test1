"""
Parse family-data-drive.xlsx → families.json
"""
from __future__ import annotations

import json
import re
import time
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "family-data-drive.xlsx"
OUT = ROOT / "family-map" / "public" / "families.json"
GEOCODE_CACHE = ROOT / "family-map" / "public" / "geocode-cache.json"

SPOUSE_RE = re.compile(r"\(&|&\s|\bwife\b|\bhusband\b", re.IGNORECASE)
BEN_BAS_RE = re.compile(r"\b(ben|bas)\s+([A-Za-z\"']+)", re.IGNORECASE)

CITY_COORDS = {
    "akron, oh": (41.0814, -81.5190),
    "athens, oh": (39.3292, -82.1013),
    "atlanta, ga": (33.7490, -84.3880),
    "bahia blanca, argentina": (-38.7183, -62.2663),
    "budapest, hungary": (47.4979, 19.0402),
    "chagrin falls, oh": (41.4362, -81.3865),
    "cleveland, oh": (41.4993, -81.6944),
    "cleveland, ohio": (41.4993, -81.6944),
    "cracow, poland": (50.0647, 19.9450),
    "crown heights": (40.6690, -73.9442),
    "cle/ny": (40.7128, -74.0060),
    "downtown cleveland": (41.4993, -81.6944),
    "kent, oh": (41.1537, -81.3579),
    "lod, israel": (31.9510, 34.8881),
    "london, england": (51.5074, -0.1278),
    "lyndhurst, ohio": (41.5201, -81.4887),
    "ma'ale ephraim, israel": (32.0722, 35.3947),
    "migdal ha'amek, israel": (32.6764, 35.2395),
    "munich, germany": (48.1351, 11.5820),
    "new york city": (40.7128, -74.0060),
    "ningbo, china": (29.8683, 121.5440),
    "ofarim, israel": (32.0561, 34.9336),
    "ora-aminadav, israel": (31.7500, 35.1500),
    "parkland, fl": (26.3100, -80.2370),
    "phuket, thailand": (7.8804, 98.3923),
    "ramot hashavim, israel": (32.1656, 34.9322),
    "shanghai, china": (31.2304, 121.4737),
    "solon, oh": (41.3898, -81.4412),
    "thane, india": (19.2183, 72.9781),
    "the bronx": (40.8448, -73.8648),
    "twinsburg, oh": (41.3126, -81.4401),
    "tzur itzchak, israel": (32.2167, 34.9833),
    "winter park, fl": (28.6000, -81.3392),
    "winterpark, fl": (28.6000, -81.3392),
    "university heights, oh": (41.4947, -81.5188),
}

_geocode_cache: dict = {}


def fmt_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def parse_coords(value) -> tuple[float, float] | None:
    if not value:
        return None
    text = str(value).strip().strip("'")
    if "," not in text:
        return None
    a, b = text.split(",", 1)
    try:
        lat, lng = float(a.strip()), float(b.strip())
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return lat, lng
    except ValueError:
        pass
    return None


def is_spouse(name: str) -> bool:
    return bool(SPOUSE_RE.search(name))


def first_name(name: str) -> str:
    return re.split(r"[,()]", name)[0].strip().split()[-1].lower()


def is_child_of(name: str, parents: list[str]) -> bool:
    m = BEN_BAS_RE.search(name)
    if not m:
        return False
    ref = m.group(2).lower().strip("'\"")
    return any(ref in p.lower() or first_name(p) == ref for p in parents)


def couple_label(parents: list[dict]) -> str:
    if len(parents) >= 2:
        return f"{parents[0]['name']} & {parents[1]['name']}"
    return parents[0]["name"] if parents else "משפחה"


def load_cache() -> dict:
    if GEOCODE_CACHE.exists():
        return json.loads(GEOCODE_CACHE.read_text(encoding="utf-8"))
    return {}


def geocode(addr: str) -> tuple[float, float] | None:
    key = addr.strip().lower()
    if not key:
        return None
    if key in _geocode_cache:
        e = _geocode_cache[key]
        return e["lat"], e["lng"]
    for ck, coords in CITY_COORDS.items():
        if ck in key or key in ck:
            _geocode_cache[key] = {"lat": coords[0], "lng": coords[1]}
            return coords
    try:
        from geopy.geocoders import Nominatim

        time.sleep(1.05)
        r = Nominatim(user_agent="family-world-map/2.0").geocode(addr)
        if r:
            _geocode_cache[key] = {"lat": r.latitude, "lng": r.longitude}
            return r.latitude, r.longitude
    except Exception as e:
        print(f"  geocode fail: {addr[:50]} — {e}")
    return None


def resolve_coords(rows: list[dict]) -> tuple[float, float, str] | None:
    lat = lng = None
    addr = ""
    for r in rows:
        if r.get("lat") is not None:
            lat, lng = r["lat"], r["lng"]
        if r.get("address"):
            addr = r["address"]
    if lat is None and addr:
        c = geocode(addr)
        if c:
            lat, lng = c
    if lat is None:
        return None
    return lat, lng, addr or f"{lat:.4f}, {lng:.4f}"


def read_row(ws, row: int) -> dict | None:
    name = ws.cell(row, 2).value
    if not name:
        return None
    c = parse_coords(ws.cell(row, 4).value)
    return {
        "name": str(name).strip(),
        "jewishBirthday": str(ws.cell(row, 1).value or "").strip(),
        "birthday": fmt_date(ws.cell(row, 3).value),
        "jewishDateShort": str(ws.cell(row, 6).value or "").strip(),
        "address": str(ws.cell(row, 5).value or "").strip(),
        "lat": c[0] if c else None,
        "lng": c[1] if c else None,
        "row": row,
    }


def is_likely_child(name: str) -> bool:
    lower = name.lower()
    return bool(BEN_BAS_RE.search(name)) or "baby" in lower


def rows_to_household(rows: list[dict], generation: int) -> dict | None:
    if not rows:
        return None

    if generation == 1:
        n = min(2, len(rows))
        parents = [{**rows[i], "role": "parent"} for i in range(n)]
        children = [{**rows[i], "role": "child"} for i in range(n, len(rows))]
    else:
        parents = [{**rows[0], "role": "parent"}]
        if len(rows) > 1 and not is_likely_child(rows[1]["name"]):
            parents.append({**rows[1], "role": "parent"})
            child_start = 2
        else:
            child_start = 1
        children = [{**rows[j], "role": "child"} for j in range(child_start, len(rows))]

    loc = resolve_coords(rows)
    if not loc:
        return None
    lat, lng, address = loc
    members = []
    for m in parents + children:
        members.append({**m, "lat": lat, "lng": lng, "address": address})

    hid = re.sub(r"[^a-zA-Z0-9\-]", "", f"g{generation}-{lat:.3f}-{parents[0]['name'][:10]}")
    return {
        "id": hid,
        "generation": generation,
        "label": couple_label(parents),
        "address": address,
        "lat": lat,
        "lng": lng,
        "parents": [{**p, "lat": lat, "lng": lng, "address": address} for p in parents],
        "childrenAtHome": [
            {**c, "lat": lat, "lng": lng, "address": address} for c in children
        ],
        "gen3Families": [],
        "members": members,
    }


def collect_records(ws) -> list[dict]:
    records = []
    empty = 0
    for row in range(3, ws.max_row + 1):
        p = read_row(ws, row)
        if not p:
            empty += 1
            continue
        if p["name"].startswith("ABC Order"):
            continue
        records.append({**p, "gap": empty})
        empty = 0
    return records


def split_gen1_trees(records: list[dict]) -> list[list[dict]]:
    trees, current = [], []
    for rec in records:
        if rec["gap"] >= 2 and current:
            trees.append(current)
            current = [rec]
        elif rec["gap"] >= 2:
            current = [rec]
        else:
            current.append(rec)
    if current:
        trees.append(current)
    return trees


def parse_gen1_tree(records: list[dict]) -> dict | None:
    """Split one gen-1 block into couple + gen-2/gen-3 families."""
    if not records:
        return None

    # Gen-1 couple: from tree start until first gap==1 (after row 0)
    gen1_rows = [records[0]]
    i = 1
    while i < len(records) and records[i]["gap"] < 1:
        gen1_rows.append(records[i])
        i += 1

    # Remaining segments split by gap==1
    segments: list[list[dict]] = []
    current: list[dict] = []
    while i < len(records):
        rec = records[i]
        if rec["gap"] >= 1 and current:
            segments.append(current)
            current = [rec]
        elif rec["gap"] >= 1:
            current = [rec]
        else:
            current.append(rec)
        i += 1
    if current:
        segments.append(current)

    gen2_families = []
    for seg in segments:
        parent_names = [seg[0]["name"]]
        if len(seg) > 1:
            parent_names.append(seg[1]["name"])
        # Sub-split segment by internal gap==1 into gen3 units
        sub_units: list[list[dict]] = []
        sub: list[dict] = []
        for j, rec in enumerate(seg):
            if j > 0 and rec["gap"] == 1:
                sub_units.append(sub)
                sub = [rec]
            else:
                sub.append(rec)
        if sub:
            sub_units.append(sub)

        if not sub_units:
            continue
        hh2 = rows_to_household(sub_units[0], 2)
        if not hh2:
            continue
        gen3_list = []
        g2_parent_names = [p["name"] for p in hh2["parents"]]
        for u in sub_units[1:]:
            gen = 3 if is_child_of(u[0]["name"], g2_parent_names) else 2
            hh = rows_to_household(u, gen)
            if gen == 3 and hh:
                gen3_list.append(hh)
            elif gen == 2 and hh:
                # mis-split: treat as sibling gen2
                gen2_families.append(hh)
        hh2["gen3Families"] = gen3_list
        gen2_families.append(hh2)

    hh1 = rows_to_household(gen1_rows, 1)
    return {
        "id": (hh1 or {}).get("id", "tree") + "-root",
        "label": (hh1 or {}).get("label", records[0]["name"]),
        "generation": 1,
        "household": hh1,
        "families": gen2_families,
    }


def build_locations(trees: list[dict]) -> list[dict]:
    by_key: dict[str, dict] = {}

    def add(hh, tree_label, path):
        if not hh or hh.get("lat") is None:
            return
        key = f"{hh['lat']:.5f},{hh['lng']:.5f}"
        loc = by_key.setdefault(
            key,
            {
                "id": key.replace(",", "-").replace(".", "_"),
                "lat": hh["lat"],
                "lng": hh["lng"],
                "address": hh["address"],
                "name": hh["address"],
                "households": [],
            },
        )
        loc["households"].append({**hh, "treeLabel": tree_label, "path": path})

    for tree in trees:
        tl = tree["label"]
        if tree.get("household"):
            add(tree["household"], tl, tl)
        for g2 in tree.get("families", []):
            add(g2, tl, f"{tl} › {g2['label']}")
            for g3 in g2.get("gen3Families", []):
                add(g3, tl, f"{tl} › {g2['label']} › {g3['label']}")

    locs = list(by_key.values())
    for loc in locs:
        loc["familyCount"] = len(loc["households"])
        loc["memberCount"] = sum(len(h["members"]) for h in loc["households"])
    return sorted(locs, key=lambda x: x["name"])


def count_members(trees) -> int:
    t = 0
    for tree in trees:
        if tree.get("household"):
            t += len(tree["household"]["members"])
        for g2 in tree.get("families", []):
            t += len(g2["members"])
            for g3 in g2.get("gen3Families", []):
                t += len(g3["members"])
    return t


def main():
    global _geocode_cache
    _geocode_cache = load_cache()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["All Families"]
    records = collect_records(ws)
    trees = [t for r in split_gen1_trees(records) if (t := parse_gen1_tree(r))]
    GEOCODE_CACHE.write_text(
        json.dumps(_geocode_cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    locations = build_locations(trees)
    payload = {
        "generatedAt": datetime.now().isoformat(),
        "locationCount": len(locations),
        "treeCount": len(trees),
        "memberCount": count_members(trees),
        "trees": trees,
        "locations": locations,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}: {len(trees)} trees, {len(locations)} locations, {payload['memberCount']} members")


if __name__ == "__main__":
    main()
